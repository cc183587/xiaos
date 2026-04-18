import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 批次详细数据
data = [
    ['3月7日', 'X3', '-', 1000, 0.63, 630.00, '已出库'],
    ['3月7日', '蓝牙带灯', '-', 200, 1.20, 240.00, '已出库'],
    ['3月8日', 'X3', '-', 187, 0.47, 87.89, '已出库'],
    ['3月12日', 'F9 头弓', '粉色', 294, 0.80, 235.20, '已出库'],
    ['3月16日', 'F9 头弓', '黑色', 170, 0.80, 136.00, '已出库'],
    ['3月17日', 'F9 头弓', '黑色', 128, 0.80, 102.40, '已出库'],
    ['3月18日', 'F9 头弓', '米白色', 200, 0.80, 160.00, '已出库'],
    ['3月19日', 'F9 头弓', '米白色', 300, 0.80, 240.00, '已出库'],
    ['3月21日', 'F9 头弓', '米白色', 300, 0.80, 240.00, '已出库'],
    ['3月22日', 'F9 头弓', '米白色', 195, 0.80, 156.00, '已出库'],
    ['3月24日', 'F9 头弓', '粉色', 404, 0.50, 202.00, '已出库'],
    ['3月29日', 'P9 头弓', '黑色', 486, 0.50, 243.00, '已出库'],
]

# 核对计算
total_qty = sum([d[3] for d in data])  # 3864
total_cost = sum([d[5] for d in data])  # 2672.49

df_batches = pd.DataFrame(data, columns=['日期', '产品', '颜色', '数量', '单价', '成本', '状态'])

# 产品汇总数据 (按实际数据计算)
summary_data = [
    ['X3', 1187, 717.89],
    ['蓝牙带灯', 200, 240.00],
    ['F9 头弓', 1991, 1391.60],
    ['P9 头弓', 486, 243.00],
]
df_summary = pd.DataFrame(summary_data, columns=['产品', '总数量', '总成本'])

# 创建Excel文件
wb = Workbook()

# 第一个工作表：批次明细
ws1 = wb.active
ws1.title = '批次明细'

# 标题
ws1['A1'] = '批次数据查看 - 2025年3月'
ws1['A1'].font = Font(size=16, bold=True, color='C0392B')
ws1.merge_cells('A1:G1')
ws1['A1'].alignment = Alignment(horizontal='center')

# 汇总信息
ws1['A3'] = '总批次数量：12 个'
ws1['A4'] = '总入库数量：3,864 件'
ws1['A5'] = '总出库数量：3,864 件'
ws1['A6'] = '总成本：¥2,672.49'
for row in range(3, 7):
    ws1[f'A{row}'].font = Font(bold=True)

# 表头
headers = ['日期', '产品', '颜色', '数量', '单价', '成本', '状态']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=8, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# 数据行
for row_idx, row_data in enumerate(data, 9):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center' if col_idx in [1, 2, 3, 7] else 'right')
        if col_idx in [5, 6]:  # 单价和成本列
            cell.number_format = '¥#,##0.00'

# 设置列宽
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 10

# 第二个工作表：产品汇总
ws2 = wb.create_sheet('产品汇总')

# 标题
ws2['A1'] = '按产品汇总'
ws2['A1'].font = Font(size=14, bold=True)
ws2.merge_cells('A1:C1')
ws2['A1'].alignment = Alignment(horizontal='center')

# 表头
summary_headers = ['产品', '总数量', '总成本']
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# 数据行
for row_idx, row_data in enumerate(summary_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center' if col_idx == 1 else 'right')
        if col_idx == 3:  # 总成本列
            cell.number_format = '¥#,##0.00'

# 设置列宽
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12

# 保存文件
output_path = 'C:/Users/J.s/WorkBuddy/20260323021915/批次数据查看-2025年3月.xlsx'
wb.save(output_path)
print(f'已导出到: {output_path}')
